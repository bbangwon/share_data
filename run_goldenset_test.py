#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import sys
import json
import argparse
import datetime
import time
import httpx

# pandas 및 openpyxl 의존성 검사
try:
    import pandas as pd
except ImportError:
    print("🚨 [Error] 'pandas' 라이브러리가 설치되어 있지 않습니다.")
    print("테스트 실행 전에 다음 명령어로 의존성을 설치해 주세요:")
    print("  pip install pandas openpyxl")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("🚨 [Error] 'openpyxl' 라이브러리가 설치되어 있지 않습니다.")
    print("테스트 실행 전에 다음 명령어로 의존성을 설치해 주세요:")
    print("  pip install openpyxl")
    sys.exit(1)


def parse_arguments():
    parser = argparse.ArgumentParser(description="Busan AI RAG Goldenset CLI Test App")
    
    # 기본 경로 설정
    default_goldenset = os.path.join(os.path.dirname(__file__), "goldenset.json")
    default_output_dir = os.path.join(os.path.dirname(__file__), "reports")
    
    parser.add_argument(
        "--goldenset-path",
        type=str,
        default=default_goldenset,
        help=f"골든셋 JSON 파일 경로 (기본값: {default_goldenset})"
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=30,
        help="테스트를 진행할 질문의 최대 개수 (기본값: 30)"
    )
    parser.add_argument(
        "--graph-url",
        type=str,
        default="http://99.1.82.168:38080",
        help="Graph-WAS 서버 주소 (기본값: http://99.1.82.168:38080)"
    )
    parser.add_argument(
        "--api-url",
        type=str,
        default="http://99.1.82.207:8080",
        help="LLM-Studio-API-Tester 서버 주소 (기본값: http://99.1.82.207:8080)"
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=1.0,
        help="각 질문 테스트 사이의 대기 시간(초) (기본값: 1.0)"
    )
    parser.add_argument(
        "--output-dir",
        type=str,
        default=default_output_dir,
        help=f"엑셀 리포트가 저장될 디렉토리 (기본값: {default_output_dir})"
    )
    
    return parser.parse_args()


def load_goldenset(file_path: str, limit: int):
    if not os.path.exists(file_path):
        print(f"❌ 골든셋 파일을 찾을 수 없습니다: {file_path}")
        sys.exit(1)
        
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            
        if not isinstance(data, list):
            print("❌ 골든셋 파일의 루트 노드는 JSON 리스트 형태여야 합니다.")
            sys.exit(1)
            
        print(f"ℹ️ 골든셋 파일 로드 완료. 총 {len(data)}개의 데이터 중 상위 {limit}개를 테스트합니다.")
        return data[:limit]
    except Exception as e:
        print(f"❌ 골든셋 파일 로드 중 오류 발생: {e}")
        sys.exit(1)


def main():
    args = parse_arguments()
    
    # 1. 골든셋 로드
    goldenset_data = load_goldenset(args.goldenset_path, args.limit)
    
    # 2. 테스트 결과 리스트 초기화
    results = []
    
    print("\n==================================================")
    print("🚀 부산 AI RAG 골든셋 자동 테스트 시작")
    print(f"🔗 Graph-WAS 주소: {args.graph_url}")
    print(f"🔗 LLM-Studio API 주소: {args.api_url}")
    print(f"⏳ 테스트 딜레이: {args.delay}초")
    print("==================================================\n")
    
    success_count = 0
    fail_count = 0
    start_time_all = time.time()
    
    # 3. 루프 실행
    for idx, item in enumerate(goldenset_data, 1):
        # '정제 질의문'이 없으면 일반 '질의문' 사용
        user_query = item.get("정제 질의문") or item.get("질의문") or ""
        user_query = user_query.strip()
        
        if not user_query:
            print(f"⚠️ [{idx}/{len(goldenset_data)}] 스킵: 유효한 질의문이 없습니다.")
            continue
            
        print(f"📝 [{idx}/{len(goldenset_data)}] 질문: \"{user_query}\"")
        
        # 기본 기록 정보
        graph_was_params = {"query": user_query}
        graph_was_response = {}
        llm_studio_payload = {}
        llm_studio_response = {}
        final_answer = ""
        status = "SUCCESS"
        error_msg = ""
        
        duration_graph_ms = 0.0
        duration_llm_ms = 0.0
        
        try:
            # --- 1단계: Graph-WAS API 호출 ---
            print("  └─ 🔍 [1/2] Graph-WAS 검색 요청 중...")
            t_graph_start = time.time()
            
            with httpx.Client(timeout=30.0) as client:
                graph_res = client.post(
                    f"{args.graph_url}/search",
                    json=graph_was_params
                )
                graph_res.raise_for_status()
                graph_was_response = graph_res.json()
                
            t_graph_end = time.time()
            duration_graph_ms = (t_graph_end - t_graph_start) * 1000.0
            
            search_result = graph_was_response.get("search_result", [])
            prompt = graph_was_response.get("prompt", "")
            
            # --- 2단계: LLM-Studio-API-Tester 호출 ---
            print("  └─ ⚡ [2/2] LLM-Studio 최종 답변 생성 중...")
            
            rag_psgs = []
            for doc in search_result:
                rag_psgs.append({
                    "id": doc.get("id") or "",
                    "passage": doc.get("context_text") or ""
                })
                
            current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            llm_studio_payload = {
                "user_query": prompt if prompt else user_query,
                "datetime": current_time,
                "stream": False,
                "dialog_history": [],
                "is_rag": True if rag_psgs else False,
                "rag_psgs": rag_psgs
            }
            
            t_llm_start = time.time()
            
            with httpx.Client(timeout=90.0) as client:
                llm_res = client.post(
                    f"{args.api_url}/api/llm/qa-response",
                    json=llm_studio_payload
                )
                llm_res.raise_for_status()
                llm_studio_response = llm_res.json()
                
            t_llm_end = time.time()
            duration_llm_ms = (t_llm_end - t_llm_start) * 1000.0
            
            # 응답 코드 체크
            response_code = llm_studio_response.get("response_code")
            if response_code and response_code != "C20000":
                status = "FAIL"
                error_msg = llm_studio_response.get("response_message") or f"상위 서버 에러 코드: {response_code}"
                final_answer = f"🚨 오류 발생: {error_msg}"
                fail_count += 1
                print(f"  ❌ 에러 발생: {error_msg}")
            else:
                final_answer = llm_studio_response.get("llm_result", {}).get("answer", "")
                success_count += 1
                print(f"  ✅ 완료 (소요시간: Graph={duration_graph_ms:.1f}ms, LLM={duration_llm_ms:.1f}ms)")
                
        except Exception as e:
            status = "FAIL"
            error_msg = str(e)
            final_answer = f"🚨 통신 오류: {error_msg}"
            fail_count += 1
            print(f"  ❌ 호출 실패: {error_msg}")
            
        # 결과 추가
        results.append({
            "No": idx,
            "사용자 질의문": user_query,
            "Graph-WAS 입력 파라미터": json.dumps(graph_was_params, ensure_ascii=False, indent=2),
            "Graph-WAS 응답값": json.dumps(graph_was_response, ensure_ascii=False, indent=2),
            "LLM-Studio 입력 파라미터": json.dumps(llm_studio_payload, ensure_ascii=False, indent=2),
            "LLM-Studio 응답값": json.dumps(llm_studio_response, ensure_ascii=False, indent=2),
            "최종 답변": final_answer,
            "상태": status,
            "에러 메시지": error_msg,
            "Graph-WAS 소요시간(ms)": round(duration_graph_ms, 2),
            "LLM-Studio 소요시간(ms)": round(duration_llm_ms, 2),
            "총 소요시간(ms)": round(duration_graph_ms + duration_llm_ms, 2)
        })
        
        # 딜레이 대기
        if idx < len(goldenset_data) and args.delay > 0:
            time.sleep(args.delay)
            
    # 4. 결과 DataFrame 생성 및 엑셀 저장
    df = pd.DataFrame(results)
    
    # 출력 폴더 생성
    os.makedirs(args.output_dir, exist_ok=True)
    
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"test_report_{timestamp}.xlsx"
    output_path = os.path.join(args.output_dir, file_name)
    
    print("\n💾 엑셀 파일 저장 중...")
    try:
        # Excel 파일 저장 및 스타일링
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="테스트 결과")
            
            # 셀 스타일링을 위한 openpyxl 객체 접근
            workbook = writer.book
            worksheet = writer.sheets["테스트 결과"]
            
            # 컬럼 너비 자동 조정
            for col in worksheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    # 줄바꿈이 있는 텍스트는 가장 긴 라인을 기준으로 계산
                    val_str = str(cell.value or "")
                    lines = val_str.split("\n")
                    for line in lines:
                        if len(line) > max_len:
                            max_len = len(line)
                # 너무 넓어지지 않도록 최대 너비 제한 (예: 60)
                worksheet.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 60)
                
            # 정렬 및 줄바꿈 허용 (내용이 긴 열은 정렬 방식을 줘서 엑셀에서 보기 편하도록 함)
            from openpyxl.styles import Alignment, PatternFill, PatternFill, Font
            
            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
            
            # 헤더 스타일 적용
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
            # 바디 텍스트 좌측 정렬 및 자동 줄바꿈
            for row in worksheet.iter_rows(min_row=2, max_row=len(results) + 1, min_col=1, max_col=12):
                for cell in row:
                    # No, 상태, 시간 관련 컬럼은 가운데 정렬
                    col_idx = cell.column
                    if col_idx in [1, 8, 10, 11, 12]:  # No, 상태, 소요시간들
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    cell.font = Font(name="맑은 고딕", size=10)
                    
        print(f"🎉 테스트 리포트 저장 완료! -> [보고서 파일]({output_path})")
    except Exception as e:
        print(f"❌ 엑셀 저장 실패: {e}")
        
    # 최종 결과 요약 출력
    duration_all = time.time() - start_time_all
    print("\n==================================================")
    print("📊 테스트 결과 요약")
    print(f"⏱️ 총 소요 시간: {duration_all:.2f}초")
    print(f"✅ 성공: {success_count}건")
    print(f"❌ 실패: {fail_count}건")
    print(f"📈 성공률: {(success_count / len(goldenset_data)) * 100:.1f}%")
    print("==================================================\n")


if __name__ == "__main__":
    main()
