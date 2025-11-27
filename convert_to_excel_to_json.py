import json
import sys

import openpyxl


def convert_excel_to_json(input_filename, output_filename):
    try:
        # 1. 엑셀 파일 로드 (data_only=True는 수식이 아닌 값을 가져옴)
        wb = openpyxl.load_workbook(input_filename, data_only=True)
        ws = wb.active  # 현재 활성화된 시트 선택

        # 2. '질의문' 헤더 찾기
        # 몇 번째 행, 몇 번째 열에 '질의문'이 있는지 찾습니다.
        header_row_index = -1
        question_col_index = -1
        check_point_index = -1
        ref_doc_index = -1
        ref_doc_page_index = -1

        for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            # 행(row) 안에 '질의문'이라는 텍스트가 있는지 확인
            if "질의문" in row:
                header_row_index = r_idx
                question_col_index = row.index("질의문")  # 0부터 시작하는 인덱스
                check_point_index = row.index("체크사항")
                ref_doc_index = row.index("답변 참조 문서명")
                ref_doc_page_index = row.index("참조 페이지")
                break

        if header_row_index == -1:
            print("오류: 엑셀 파일에서 '질의문' 헤더를 찾을 수 없습니다.")
            return

        # 3. 데이터 변환
        formatted_data = []

        # 헤더 다음 행부터 데이터를 읽어옵니다.
        for row in ws.iter_rows(min_row=header_row_index + 1, values_only=True):
            if row[check_point_index] != "참조문서확정":
                continue

            question_text = row[question_col_index]
            ref_doc_text = row[ref_doc_index]
            ref_doc_page_text = row[ref_doc_page_index]

            # 내용이 없으면 건너뜀
            if question_text is None or str(question_text).strip() == "":
                continue

            # (1) 시작 메시지 객체
            start_obj = {"text": "시작", "postback": "start"}

            # (2) 질문 객체
            question_obj = {
                "text": str(question_text),
                "ref_doc": str(ref_doc_text),
                "ref_page": int(str(ref_doc_page_text)),
            }

            formatted_data.append(start_obj)
            formatted_data.append(question_obj)

        # 4. JSON 파일 저장
        with open(output_filename, "w", encoding="utf-8") as f:
            json.dump(formatted_data, f, ensure_ascii=False, indent=2)

        print(f"변환 완료! '{output_filename}' 파일이 생성되었습니다.")
        print(f"총 {len(formatted_data) // 2}개의 질문이 처리되었습니다.")

    except FileNotFoundError:
        print(f"오류: 입력 파일 '{input_filename}'을(를) 찾을 수 없습니다.")
    except Exception as e:
        print(f"오류 발생: {e}")


if __name__ == "__main__":
    # 실행 시 인자 개수 확인 (스크립트명 + 입력파일 + 출력파일 = 3개)
    if len(sys.argv) != 3:
        print("사용법: python 스크립트명.py <입력_엑셀파일.xlsx> <출력_파일.json>")
        print("예시: python converter.py input.xlsx output.json")
    else:
        input_file = sys.argv[1]
        output_file = sys.argv[2]
        convert_excel_to_json(input_file, output_file)
