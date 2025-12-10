import os
import sys

import openpyxl
from openpyxl.styles import Alignment, PatternFill


def normalize_val(val):
    """값을 문자열로 변환하고, 정수형 실수는 정수 문자로 변환 (예: 1.0 -> '1')"""
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def parse_pages(page_str):
    """
    페이지 문자열을 파싱하여 개별 페이지 번호의 리스트(문자열)를 반환
    예: "1, 3~5" -> {'1', '3', '4', '5'}
    """
    pages = set()
    s = str(page_str).strip()
    if not s or s.lower() == "nan" or s == "none":
        return pages

    # 쉼표로 분리
    parts = s.split(",")
    for part in parts:
        part = part.strip()
        if "~" in part:
            # 범위 처리 (예: 1~3)
            try:
                start_s, end_s = part.split("~")
                start = int(start_s)
                end = int(end_s)
                for i in range(start, end + 1):
                    pages.add(str(i))
            except ValueError:
                pages.add(part)
        else:
            # 단일 페이지
            pages.add(part)
    return pages


def check_red(cell):
    # 글자색이 있고 RGB값이 빨간색(FFFF0000 등)인지 확인
    if cell.font and cell.font.color and hasattr(cell.font.color, "rgb"):
        color = str(cell.font.color.rgb)
        # Alpha channel 포함(FFFF0000) 또는 미포함(FF0000/00FF0000) 케이스 대응
        if "FF0000" in color and color != "00000000":
            if color == "FFFF0000" or color == "00FF0000":
                return True
    return False


def process_excel_log(input_file, output_file, condition_mode="AND"):
    print(f"[{input_file}] 파일을 읽는 중... (조건 모드: {condition_mode})")

    # 1. 엑셀 파일 불러오기
    try:
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active
    except FileNotFoundError:
        print(f"오류: 입력 파일 '{input_file}'을(를) 찾을 수 없습니다.")
        sys.exit(1)
    except Exception as e:
        print(f"오류 발생: {e}")
        sys.exit(1)

    # 3. D열, E열 신규 삽입 (4번째 위치에 2개 열)
    ws.insert_cols(5, 2)

    # 헤더 설정
    ws["E1"] = "검색TOP5포함"
    ws["F1"] = "SLM정답청크선택"

    widths = [5, 30, 20, 10, 10, 10, 10, 50, 10, 10, 10, 30, 10]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    print("데이터 분석 및 처리 중...")

    # 배경색 스타일 정의
    light_green_fill = PatternFill(
        start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"
    )

    # 5개씩 처리
    max_row = ws.max_row
    for row_start in range(2, max_row + 1, 5):
        row_end = min(row_start + 4, max_row)

        # 병합 처리
        ws.merge_cells(
            start_row=row_start, start_column=7, end_row=row_end, end_column=7
        )
        ws.merge_cells(
            start_row=row_start, start_column=8, end_row=row_end, end_column=8
        )

        cell_c = ws.cell(row=row_start, column=3)  # 답변참조문서
        cell_d = ws.cell(row=row_start, column=4)  # 참조페이지
        cell_e = ws.cell(row=row_start, column=5)  # 참조문서검색 (Target)
        cell_f = ws.cell(row=row_start, column=6)  # SLM참조여부

        cell_e.alignment = Alignment(vertical="top")
        cell_f.alignment = Alignment(vertical="top")

        c_val = normalize_val(cell_c.value)
        d_pages = parse_pages(cell_d.value)

        # 5개씩 참조 데이터베이스 구축
        ref_db = dict()
        ref_db_color = dict()

        for r_item in range(row_start, row_end + 1):
            cell_l = ws.cell(row=r_item, column=12)
            cell_m = ws.cell(row=r_item, column=13)

            file_val = normalize_val(cell_l.value)
            page_val = normalize_val(cell_m.value)

            if file_val and page_val:
                ref_db[(file_val, page_val)] = r_item
                if check_red(cell_l) and check_red(cell_m):
                    ref_db_color[(file_val, page_val)] = r_item

        # 검증 로직
        if c_val and d_pages:
            found_count = 0
            referenced_count = 0
            total_pages = len(d_pages)

            for page in d_pages:
                key = (c_val, page)

                # 1. DB에 존재하는지 확인
                if key in ref_db:
                    found_count += 1
                    # 매칭된 행 색칠하기
                    r_item = ref_db[key]
                    for col in range(9, 14):  # 8~12열 (H~L)
                        ws.cell(row=r_item, column=col).fill = light_green_fill

                # 2. 빨간색(SLM참조)인지 확인
                if key in ref_db_color:
                    referenced_count += 1

            # 조건 모드에 따른 결과 판정
            if condition_mode == "OR":
                # 하나라도 존재하면 True
                exist_bool = found_count > 0
                ref_bool = referenced_count > 0
            else:
                # AND (기본값): 모든 페이지가 존재해야 True
                exist_bool = found_count == total_pages
                ref_bool = referenced_count == total_pages

            cell_e.value = "포함" if exist_bool else "미포함"
            cell_f.value = "선택" if ref_bool else "미선택"

        else:
            # 참조 문서나 페이지 정보가 없는 경우
            cell_e.value = "미포함"
            cell_f.value = "미선택"

    # 워크시트 추가
    try:
        statistics = wb.create_sheet(title="Statistics")
        wb._sheets.insert(0, wb._sheets.pop())

        statistics["A1"] = "총 질문 수"
        statistics["B1"] = "=COUNTA(Logs!A:A) - 1"
        statistics["A2"] = "검색TOP5포함"
        statistics["B2"] = '=COUNTIF(Logs!E:E, "포함")'
        statistics["C2"] = "=B2/B1"
        statistics["A3"] = "SLM정답청크선택"
        statistics["B3"] = '=COUNTIF(Logs!F:F, "선택")'
        statistics["C3"] = "=B3/B1"

        # C2, C3를 %로 표시
        statistics["C2"].number_format = "0.00%"
        statistics["C3"].number_format = "0.00%"

        statistics.column_dimensions["A"].width = 20
    except Exception as e:
        print(f"워크시트 추가 중 오류 발생: {e}")

    # 5. 저장
    try:
        wb.save(output_file)
        print(f"완료! 결과가 '{output_file}'(으)로 저장되었습니다.")
    except Exception as e:
        print(f"저장 중 오류 발생: {e}")


# --- 메인 실행 블록 ---
if __name__ == "__main__":
    # 인자 개수 확인
    # 기본: script.py input (len=2)
    # 옵션: script.py input AND (len=3)
    if len(sys.argv) < 2:
        print("사용법: python 스크립트명.py <입력파일경로> [조건옵션:AND|OR]")
        print("예시 (기본 OR): python convert.py input.xlsx")
        print("예시 (AND 모드): python convert.py input.xlsx AND")
    else:
        input_path = sys.argv[1]

        # 출력 파일명 자동 생성
        filename, ext = os.path.splitext(input_path)
        output_path = f"{filename}_correct{ext}"

        # 조건 옵션 파싱 (기본값 OR)
        mode_arg = "OR"
        if len(sys.argv) >= 3:
            # 입력받은 값을 대문자로 변환하여 확인
            if sys.argv[2].upper() == "AND":
                mode_arg = "AND"

        process_excel_log(input_path, output_path, mode_arg)
