import argparse
import json
import os

import openpyxl
from openpyxl.styles import Alignment


def load_json_data(json_path):
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            result = [item for item in data if "number" in item]
        return result
    except Exception as e:
        print(f"JSON 로드 중 오류 발생: {e}")
        return None


def detect_merge_rows_from_excel(excel_path):
    """엑셀 파일의 '질의문' 헤더가 있는 컬럼에서 머지된 행의 개수를 감지합니다."""
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        # 상위 10개 행에서 '질의문' 헤더 찾기
        q_column_idx = None
        for row in ws.iter_rows(min_row=1, max_row=10):
            for col_idx, cell in enumerate(row, start=1):
                if cell.value and str(cell.value).strip() == "질의문":
                    q_column_idx = col_idx
                    break
            if q_column_idx:
                break

        if q_column_idx is None:
            print("경고: '질의문' 헤더를 찾을 수 없습니다. 기본값 5 사용")
            wb.close()
            return 5

        merge_rows_list = []

        # '질의문' 컬럼이 포함된 모든 머지된 범위 확인
        for merged_range in ws.merged_cells.ranges:
            # 해당 컬럼이 포함된 머지 범위인지 확인
            if merged_range.min_col <= q_column_idx <= merged_range.max_col:
                # 행의 개수 계산
                row_count = merged_range.max_row - merged_range.min_row + 1
                merge_rows_list.append(row_count)

        wb.close()

        if merge_rows_list:
            # 가장 많이 나타나는 머지 행의 개수 사용
            from collections import Counter

            most_common = Counter(merge_rows_list).most_common(1)[0][0]
            return most_common

        return 5  # 기본값
    except Exception as e:
        print(f"엑셀의 머지 행 개수 감지 중 오류: {e}")
        return 5  # 오류 시 기본값


def update_excel_xlsx(input_path, output_path, data_list, merge_rows=5):
    try:
        wb = openpyxl.load_workbook(input_path)
        ws = wb.active

        # 3. A열 신규 삽입 (1번째 위치에 1개 열)
        ws.insert_cols(1, 1)
        # 4. D열, E열 신규 삽입 (4번째 위치에 2개 열)
        ws.insert_cols(3, 2)

        # 헤더 설정
        ws["A1"] = "순번"
        ws["C1"] = "답변참조문서"
        ws["D1"] = "참조페이지"

        widths = [5, 30, 20, 10, 10, 50, 10, 10, 10, 30, 10]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # 헤더 찾기 (유연한 검색)
        headers = {}
        header_row_obj = None

        # 상위 10개 행 내에서 '질의문' 헤더 찾기
        for row in ws.iter_rows(min_row=1, max_row=10):
            row_values = [str(c.value).strip() if c.value else "" for c in row]
            if "질의문" in row_values:
                header_row_obj = row
                for idx, val in enumerate(row_values):
                    if val:
                        headers[val] = idx
                break

        if header_row_obj is None:
            print("오류: 엑셀 파일에서 '질의문' 헤더를 찾을 수 없습니다.")
            return

        # 컬럼 인덱스 식별 (다양한 이름 대응)
        q_idx = headers.get("질의문")

        num_idx = headers.get("순번")
        # '답변참조문서' 또는 '답변 참조 문서명' 등
        doc_idx = headers.get("답변참조문서")
        # '참조페이지' 또는 '참조 페이지'
        page_idx = headers.get("참조페이지")
        if q_idx is None or doc_idx is None or page_idx is None:
            print("오류: 필수 컬럼을 찾을 수 없습니다.")
            print(f"확인된 헤더: {list(headers.keys())}")
            return

        # 데이터 업데이트
        updated_count = 0
        start_row_idx = header_row_obj[0].row + 1  # 헤더 다음 줄부터

        for row in ws.iter_rows(min_row=start_row_idx):
            # 질문 가져오기
            cell_q = row[q_idx]
            question_text = str(cell_q.value).strip() if cell_q.value else ""
            if question_text:
                data = data_list[updated_count]
                # 값 업데이트
                row[num_idx].alignment = Alignment(vertical="top", wrap_text=True)
                row[num_idx].value = data["number"]

                row[doc_idx].alignment = Alignment(vertical="top", wrap_text=True)
                row[doc_idx].value = data["ref_doc"]

                row[page_idx].alignment = Alignment(vertical="top", wrap_text=True)
                row[page_idx].value = data["ref_page"]

                updated_count += 1

        max_row = ws.max_row
        for row_start in range(2, max_row + 1, merge_rows):
            row_end = min(row_start + merge_rows - 1, max_row)

            # 병합 처리
            ws.merge_cells(
                start_row=row_start, start_column=4, end_row=row_end, end_column=4
            )
            ws.merge_cells(
                start_row=row_start, start_column=5, end_row=row_end, end_column=5
            )
            ws.merge_cells(
                start_row=row_start, start_column=6, end_row=row_end, end_column=6
            )

        wb.save(output_path)
        print(f"[완료] 총 {updated_count}개의 항목을 엑셀에 업데이트했습니다.")
        print(f"저장된 파일: {output_path}")

    except Exception as e:
        print(f"XLSX 처리 중 오류: {e}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="JSON 데이터를 바탕으로 엑셀 파일을 업데이트합니다.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
예제:
  python 2.update_excel_from_json.py -j data.json -e template.xlsx -o output.xlsx -r 5
  python 2.update_excel_from_json.py -j data.json -e template.xlsx
        """,
    )

    parser.add_argument(
        "-j", "--json", required=True, dest="json_file", help="입력 JSON 파일 경로"
    )
    parser.add_argument(
        "-e", "--excel", required=True, dest="excel_file", help="입력 엑셀 파일 경로"
    )
    parser.add_argument(
        "-o",
        "--output",
        dest="output_file",
        default=None,
        help="출력 엑셀 파일 경로 (기본값: 입력파일_updated.xlsx)",
    )
    parser.add_argument(
        "-r",
        "--rows",
        dest="merge_rows",
        type=int,
        default=5,
        help="머지할 행 개수 (기본값: 5)",
    )

    args = parser.parse_args()

    json_file = args.json_file
    excel_file = args.excel_file
    merge_rows = args.merge_rows

    # 출력 파일명이 없으면 원본 파일명에 '_updated'를 붙임
    if args.output_file:
        output_file = args.output_file
    else:
        fname, ext = os.path.splitext(excel_file)
        output_file = f"{fname}_updated{ext}"

    print(f"1. JSON 데이터 로딩: {json_file}")
    mapping_data = load_json_data(json_file)

    if mapping_data:
        print(f"   - {len(mapping_data)}개의 질문 데이터 로드됨")
        print(f"2. 엑셀 파일 업데이트: {excel_file}")

        # -r 옵션이 없으면 엑셀 파일에서 머지 행 개수 자동 감지
        if args.merge_rows == 5:  # 기본값
            detected_merge_rows = detect_merge_rows_from_excel(excel_file)
            print(f"   - 머지 행 개수 자동 감지: {detected_merge_rows}")
            merge_rows = detected_merge_rows
        else:
            print(f"   - 머지 행 개수: {merge_rows}")

        ext = os.path.splitext(excel_file)[1].lower()
        if ext == ".xlsx":
            update_excel_xlsx(excel_file, output_file, mapping_data, merge_rows)
        else:
            print("지원하지 않는 파일 형식입니다. (.xlsx만 지원)")
