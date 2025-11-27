import csv
import json
import os
import sys

import openpyxl


def load_json_data(json_path):
    """
    JSON 파일을 읽어 질의문을 Key로 하는 딕셔너리를 생성합니다.
    구조: {'질문텍스트': {'ref_doc': '...', 'ref_page': '...'}}
    """
    mapping = {}
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

            for item in data:
                # 텍스트(질문)가 있는 항목만 처리
                if isinstance(item, dict) and "text" in item:
                    question = str(item["text"]).strip()

                    # ref_doc이나 ref_page 키가 있다면 업데이트 대상에 포함
                    if "ref_doc" in item or "ref_page" in item:
                        mapping[question] = {
                            "ref_doc": str(item.get("ref_doc", "")),
                            "ref_page": str(item.get("ref_page", "")),
                        }
        return mapping
    except Exception as e:
        print(f"JSON 로드 중 오류 발생: {e}")
        return None


def update_excel_xlsx(input_path, output_path, data_map):
    try:
        wb = openpyxl.load_workbook(input_path)
        ws = wb.active

        # 헤더 찾기 (유연한 검색)
        headers = {}
        header_row_obj = None

        # 상위 10개 행 내에서 '질의문' 헤더 찾기
        for row in ws.iter_rows(min_row=1, max_row=10):
            row_values = [str(c.value).strip() if c.value else "" for c in row]
            if "질의문" in row_values:
                header_row_obj = row
                for idx, val in enumerate(row_values):
                    if val:
                        headers[val] = idx
                break

        if header_row_obj is None:
            print("오류: 엑셀 파일에서 '질의문' 헤더를 찾을 수 없습니다.")
            return

        # 컬럼 인덱스 식별 (다양한 이름 대응)
        q_idx = headers.get("질의문")

        # '답변참조문서' 또는 '답변 참조 문서명' 등
        doc_idx = headers.get("답변참조문서")
        if doc_idx is None:
            doc_idx = headers.get("답변 참조 문서명")

        # '참조페이지' 또는 '참조 페이지'
        page_idx = headers.get("참조페이지")
        if page_idx is None:
            page_idx = headers.get("참조 페이지")

        if q_idx is None or doc_idx is None or page_idx is None:
            print("오류: 필수 컬럼을 찾을 수 없습니다.")
            print(f"확인된 헤더: {list(headers.keys())}")
            return

        # 데이터 업데이트
        updated_count = 0
        start_row_idx = header_row_obj[0].row + 1  # 헤더 다음 줄부터

        for row in ws.iter_rows(min_row=start_row_idx):
            # 질문 가져오기
            cell_q = row[q_idx]
            question_text = str(cell_q.value).strip() if cell_q.value else ""

            if question_text in data_map:
                new_info = data_map[question_text]

                # 값 업데이트
                row[doc_idx].value = new_info["ref_doc"]
                row[page_idx].value = new_info["ref_page"]
                updated_count += 1

        wb.save(output_path)
        print(f"[완료] 총 {updated_count}개의 항목을 엑셀에 업데이트했습니다.")
        print(f"저장된 파일: {output_path}")

    except Exception as e:
        print(f"XLSX 처리 중 오류: {e}")


def update_excel_csv(input_path, output_path, data_map):
    try:
        rows = []
        with open(input_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            headers = next(reader, None)
            if not headers:
                print("오류: CSV 파일이 비어있습니다.")
                return

            # 헤더 인덱스 찾기
            header_map = {h.strip(): i for i, h in enumerate(headers)}

            q_idx = header_map.get("질의문")
            doc_idx = header_map.get("답변참조문서") or header_map.get(
                "답변 참조 문서명"
            )
            page_idx = header_map.get("참조페이지") or header_map.get("참조 페이지")

            if q_idx is None or doc_idx is None or page_idx is None:
                print(f"오류: 필수 컬럼을 찾을 수 없습니다. (헤더: {headers})")
                return

            rows.append(headers)
            updated_count = 0

            for row in reader:
                if len(row) > q_idx:
                    question_text = str(row[q_idx]).strip()
                    if question_text in data_map:
                        new_info = data_map[question_text]
                        # 행 길이가 부족하면 늘려줌
                        while len(row) <= max(doc_idx, page_idx):
                            row.append("")

                        row[doc_idx] = new_info["ref_doc"]
                        row[page_idx] = new_info["ref_page"]
                        updated_count += 1
                rows.append(row)

        with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerows(rows)

        print(f"[완료] 총 {updated_count}개의 항목을 CSV에 업데이트했습니다.")
        print(f"저장된 파일: {output_path}")

    except Exception as e:
        print(f"CSV 처리 중 오류: {e}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(
            "사용법: python update_excel_from_json.py [JSON파일] [엑셀파일] [출력파일명(선택)]"
        )
        sys.exit(1)

    json_file = sys.argv[1]
    excel_file = sys.argv[2]

    # 출력 파일명이 없으면 원본 파일명에 '_updated'를 붙임
    if len(sys.argv) >= 4:
        output_file = sys.argv[3]
    else:
        fname, ext = os.path.splitext(excel_file)
        output_file = f"{fname}_updated{ext}"

    print(f"1. JSON 데이터 로딩: {json_file}")
    mapping_data = load_json_data(json_file)

    if mapping_data:
        print(f"   - {len(mapping_data)}개의 질문 데이터 로드됨")
        print(f"2. 엑셀 파일 업데이트: {excel_file}")

        ext = os.path.splitext(excel_file)[1].lower()
        if ext == ".xlsx":
            update_excel_xlsx(excel_file, output_file, mapping_data)
        elif ext == ".csv":
            update_excel_csv(excel_file, output_file, mapping_data)
        else:
            print("지원하지 않는 파일 형식입니다. (.xlsx, .csv 만 지원)")
