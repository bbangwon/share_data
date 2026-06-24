#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import sys
import json
import datetime
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def parse_goldenset_pages(page_str):
    if not page_str or pd.isna(page_str):
        return set()
    pages = set()
    for p in str(page_str).split(','):
        p = p.strip()
        if p.isdigit():
            pages.add(int(p))
    return pages

def extract_graph_was_pages(graph_was_response_str):
    pages = set()
    if not graph_was_response_str or pd.isna(graph_was_response_str):
        return pages
    try:
        data = json.loads(graph_was_response_str)
        search_result = data.get("search_result", [])
        for doc in search_result:
            page_val = doc.get("page")
            if isinstance(page_val, list):
                for p in page_val:
                    if isinstance(p, int):
                        pages.add(p)
                    elif isinstance(p, str) and p.isdigit():
                        pages.add(int(p))
            elif isinstance(page_val, (int, float)):
                pages.add(int(page_val))
            elif isinstance(page_val, str) and page_val.isdigit():
                pages.add(int(page_val))
    except Exception as e:
        pass
    return pages

def extract_answer_pages(final_answer):
    pages = set()
    if not final_answer or pd.isna(final_answer):
        return pages
    # p. 39, p. 128,131 등 패턴
    matches = re.findall(r'p\.\s*([0-9\s,]+)', str(final_answer))
    for match in matches:
        for part in match.split(','):
            part = part.strip()
            if part.isdigit():
                pages.add(int(part))
    return pages

def write_summary_sheet(workbook, results):
    if "결과 요약" in workbook.sheetnames:
        del workbook["결과 요약"]
        
    summary_sheet = workbook.create_sheet(title="결과 요약", index=0)
    summary_sheet.views.sheetView[0].showGridLines = True
    
    # 폰트 및 채우기 설정
    title_font = Font(name="맑은 고딕", size=16, bold=True, color="1F497D")
    subtitle_font = Font(name="맑은 고딕", size=10, italic=True, color="595959")
    header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="맑은 고딕", size=11, bold=True)
    normal_font = Font(name="맑은 고딕", size=10)
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    
    # 테두리 설정
    thin_side = Side(style='thin', color='D9D9D9')
    thick_bottom = Side(style='medium', color='1F497D')
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thick_bottom)
    
    # 타이틀
    summary_sheet["A1"] = "📊 Busan AI RAG 골든셋 테스트 결과 요약"
    summary_sheet["A1"].font = title_font
    
    summary_sheet["A2"] = f"생성 일시: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    summary_sheet["A2"].font = subtitle_font
    
    # --- 표 1: 정확도 지표 ---
    summary_sheet["A4"] = "평가 구분"
    summary_sheet["B4"] = "대상 건수"
    summary_sheet["C4"] = "성공/매칭 건수"
    summary_sheet["D4"] = "성공률 / 매칭률 (%)"
    
    for col_let in ["A", "B", "C", "D"]:
        cell = summary_sheet[f"{col_let}4"]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border
        
    total_cases = len(results)
    api_success = sum(1 for r in results if r["상태"] == "SUCCESS")
    g_match = sum(1 for r in results if r["Graph-WAS 매칭 여부"] == "Y")
    a_match = sum(1 for r in results if r["최종 답변 매칭 여부"] == "Y")
    
    metrics = [
        ("API 호출 성공률", total_cases, api_success, (api_success / total_cases * 100) if total_cases > 0 else 0),
        ("Graph-WAS 검색 성공률 (쪽번호 일치)", total_cases, g_match, (g_match / total_cases * 100) if total_cases > 0 else 0),
        ("최종 답변 참조 성공률 (쪽번호 일치)", total_cases, a_match, (a_match / total_cases * 100) if total_cases > 0 else 0),
    ]
    
    for idx, (label, tot, succ, rate) in enumerate(metrics, 5):
        summary_sheet[f"A{idx}"] = label
        summary_sheet[f"B{idx}"] = f"{tot}건"
        summary_sheet[f"C{idx}"] = f"{succ}건"
        summary_sheet[f"D{idx}"] = f"{rate:.1f}%"
        
        for col_let in ["A", "B", "C", "D"]:
            cell = summary_sheet[f"{col_let}{idx}"]
            cell.font = bold_font if col_let == "A" else normal_font
            cell.border = cell_border
            if col_let == "A":
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
    # --- 표 2: 소요 시간 지표 ---
    time_start_row = 10
    summary_sheet[f"A{time_start_row}"] = "소요 시간 구분"
    summary_sheet[f"B{time_start_row}"] = "평균 소요 시간 (ms)"
    summary_sheet[f"C{time_start_row}"] = "최대 소요 시간 (ms)"
    summary_sheet[f"D{time_start_row}"] = "최소 소요 시간 (ms)"
    
    for col_let in ["A", "B", "C", "D"]:
        cell = summary_sheet[f"{col_let}{time_start_row}"]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border
        
    def to_float(val):
        try:
            return float(val)
        except (ValueError, TypeError):
            return 0.0
            
    graph_times = [to_float(r.get("Graph-WAS 소요시간(ms)", 0)) for r in results]
    llm_times = [to_float(r.get("LLM-Studio 소요시간(ms)", 0)) for r in results]
    total_times = [to_float(r.get("총 소요시간(ms)", 0)) for r in results]
    
    avg_graph = sum(graph_times) / total_cases if total_cases > 0 else 0
    avg_llm = sum(llm_times) / total_cases if total_cases > 0 else 0
    avg_total = sum(total_times) / total_cases if total_cases > 0 else 0
    
    max_graph = max(graph_times) if graph_times else 0
    max_llm = max(llm_times) if llm_times else 0
    max_total = max(total_times) if total_times else 0
    
    min_graph = min(graph_times) if graph_times else 0
    min_llm = min(llm_times) if llm_times else 0
    min_total = min(total_times) if total_times else 0
    
    time_metrics = [
        ("Graph-WAS 검색 소요시간", avg_graph, max_graph, min_graph),
        ("LLM-Studio 답변생성 소요시간", avg_llm, max_llm, min_llm),
        ("전체 프로세스 총 소요시간", avg_total, max_total, min_total)
    ]
    
    for idx, (label, avg, mx, mn) in enumerate(time_metrics, time_start_row + 1):
        summary_sheet[f"A{idx}"] = label
        summary_sheet[f"B{idx}"] = f"{avg:.1f} ms"
        summary_sheet[f"C{idx}"] = f"{mx:.1f} ms"
        summary_sheet[f"D{idx}"] = f"{mn:.1f} ms"
        
        for col_let in ["A", "B", "C", "D"]:
            cell = summary_sheet[f"{col_let}{idx}"]
            cell.font = bold_font if col_let == "A" else normal_font
            cell.border = cell_border
            if col_let == "A":
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
    summary_sheet.column_dimensions["A"].width = 38
    summary_sheet.column_dimensions["B"].width = 18
    summary_sheet.column_dimensions["C"].width = 18
    summary_sheet.column_dimensions["D"].width = 24

def main():
    base_dir = os.path.dirname(__file__)
    goldenset_path = os.path.join(base_dir, "goldenset.json")
    report_path = os.path.join(base_dir, "test_report_20260624_142759.xlsx")
    output_path = os.path.join(base_dir, "test_report_20260624_142759_analyzed.xlsx")
    
    print(f"📖 골든셋 파일 로드 중: {goldenset_path}")
    with open(goldenset_path, "r", encoding="utf-8") as f:
        goldenset_data = json.load(f)
        
    goldenset_map = {}
    for item in goldenset_data:
        # run_goldenset_test.py와 동일한 방식으로 query 매핑 키 생성
        q = item.get("정제 질의문") or item.get("질의문") or ""
        q = q.strip()
        if q:
            goldenset_map[q] = item.get("적용 쪽번호") or ""
            
    print(f"📊 기존 엑셀 파일 로드 중: {report_path}")
    df = pd.read_excel(report_path)
    
    results = []
    for idx, row in df.iterrows():
        user_query = str(row["사용자 질의문"]).strip()
        goldenset_page_str = goldenset_map.get(user_query, "")
        
        # 1. 골든셋 정답페이지 파싱
        goldenset_pages = parse_goldenset_pages(goldenset_page_str)
        
        # 2. Graph-WAS 응답값에서 페이지 추출
        graph_was_resp_str = row["Graph-WAS 응답값"]
        graph_was_pages = extract_graph_was_pages(graph_was_resp_str)
        
        # 3. 최종 답변에서 페이지 추출
        final_answer = row["최종 답변"]
        final_answer_pages = extract_answer_pages(final_answer)
        
        # 4. 매칭 여부 판정
        graph_was_match = "Y" if (goldenset_pages & graph_was_pages) else "N"
        final_answer_match = "Y" if (goldenset_pages & final_answer_pages) else "N"
        
        # 5. 기존 행 데이터를 복사하고 새 컬럼 추가
        row_dict = row.to_dict()
        row_dict["골든셋 정답페이지"] = goldenset_page_str
        row_dict["Graph-WAS 추출페이지"] = ", ".join(map(str, sorted(graph_was_pages)))
        row_dict["최종 답변 추출페이지"] = ", ".join(map(str, sorted(final_answer_pages)))
        row_dict["Graph-WAS 매칭 여부"] = graph_was_match
        row_dict["최종 답변 매칭 여부"] = final_answer_match
        
        results.append(row_dict)
        
    df_new = pd.DataFrame(results)
    
    # 엑셀 파일 저장
    print(f"💾 결과 엑셀 저장 중: {output_path}")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_new.to_excel(writer, index=False, sheet_name="테스트 결과")
        
        workbook = writer.book
        worksheet = writer.sheets["테스트 결과"]
        
        # 결과 요약 시트 추가
        write_summary_sheet(workbook, results)
        
        # 테스트 결과 시트 스타일링 적용
        # 컬럼 너비 자동 조정
        for col in worksheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val_str = str(cell.value or "")
                lines = val_str.split("\n")
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            worksheet.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 60)
            
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
        
        # 헤더 스타일
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        # 바디 스타일 (총 17개 컬럼)
        for row_cells in worksheet.iter_rows(min_row=2, max_row=len(results) + 1, min_col=1, max_col=17):
            for cell in row_cells:
                col_idx = cell.column
                # No, 상태, 소요시간들, 쪽번호 관련 컬럼들은 가운데 정렬
                if col_idx in [1, 8, 10, 11, 12, 13, 14, 15, 16, 17]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.font = Font(name="맑은 고딕", size=10)
                
    print("🎉 변환 작업이 완료되었습니다!")

if __name__ == "__main__":
    main()
