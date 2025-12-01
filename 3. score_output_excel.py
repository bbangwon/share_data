import sys

import openpyxl
from openpyxl.styles import Alignment, PatternFill


def normalize_val(val):
    """값을 문자열로 변환하고, 정수형 실수는 정수 문자로 변환 (예: 1.0 -> '1')"""
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def parse_pages(page_str):
    """
    페이지 문자열을 파싱하여 개별 페이지 번호의 리스트(문자열)를 반환
    예: "1, 3~5" -> {'1', '3', '4', '5'}
    """
    pages = set()
    s = str(page_str).strip()
    if not s or s.lower() == "nan" or s == "none":
        return pages

    # 쉼표로 분리
    parts = s.split(",")
    for part in parts:
        part = part.strip()
        if "~" in part:
            # 범위 처리 (예: 1~3)
            try:
                start_s, end_s = part.split("~")
                start = int(start_s)
                end = int(end_s)
                for i in range(start, end + 1):
                    pages.add(str(i))
            except ValueError:
                pages.add(part)
        else:
            # 단일 페이지
            pages.add(part)
    return pages


def check_red(cell):
    # 글자색이 있고 RGB값이 빨간색(FFFF0000 등)인지 확인
    if cell.font and cell.font.color and hasattr(cell.font.color, "rgb"):
        color = str(cell.font.color.rgb)
        # Alpha channel 포함(FFFF0000) 또는 미포함(FF0000/00FF0000) 케이스 대응
        if "FF0000" in color and color != "00000000":
            if color == "FFFF0000" or color == "00FF0000":
                return True
    return False


def process_excel_log(input_file, output_file, condition_mode="AND"):
    print(f"[{input_file}] 파일을 읽는 중... (조건 모드: {condition_mode})")

    # 1. 엑셀 파일 불러오기
    try:
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active
    except FileNotFoundError:
        print(f"오류: 입력 파일 '{input_file}'을(를) 찾을 수 없습니다.")
        sys.exit(1)
    except Exception as e:
        print(f"오류 발생: {e}")
        sys.exit(1)

    # 3. D열, E열 신규 삽입 (4번째 위치에 2개 열)
    ws.insert_cols(4, 2)

    # 헤더 설정
    ws["D1"] = "참조문서검색"
    ws["E1"] = "SLM참조여부"

    widths = [30, 20, 10, 10, 10, 10, 50, 10, 10, 10, 30, 10]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    print("데이터 분석 및 처리 중...")

    # 배경색 스타일 정의
    light_green_fill = PatternFill(
        start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"
    )

    # 5개씩 처리
    max_row = ws.max_row
    for row_start in range(2, max_row + 1, 5):
        row_end = min(row_start + 4, max_row)

        # 병합 처리
        ws.merge_cells(
            start_row=row_start, start_column=6, end_row=row_end, end_column=6
        )
        ws.merge_cells(
            start_row=row_start, start_column=7, end_row=row_end, end_column=7
        )

        cell_b = ws.cell(row=row_start, column=2)  # 답변참조문서
        cell_c = ws.cell(row=row_start, column=3)  # 참조페이지
        cell_d = ws.cell(row=row_start, column=4)  # 참조문서검색 (Target)
        cell_e = ws.cell(row=row_start, column=5)  # SLM참조여부

        cell_d.alignment = Alignment(vertical="top")
        cell_e.alignment = Alignment(vertical="top")

        b_val = normalize_val(cell_b.value)
        c_pages = parse_pages(cell_c.value)

        # 5개씩 참조 데이터베이스 구축
        ref_db = dict()
        ref_db_color = dict()

        for r_item in range(row_start, row_end + 1):
            cell_k = ws.cell(row=r_item, column=11)
            cell_l = ws.cell(row=r_item, column=12)

            file_val = normalize_val(cell_k.value)
            page_val = normalize_val(cell_l.value)

            if file_val and page_val:
                ref_db[(file_val, page_val)] = r_item
                if check_red(cell_k) and check_red(cell_l):
                    ref_db_color[(file_val, page_val)] = r_item

        # 검증 로직
        if b_val and c_pages:
            found_count = 0
            referenced_count = 0
            total_pages = len(c_pages)

            for page in c_pages:
                key = (b_val, page)

                # 1. DB에 존재하는지 확인
                if key in ref_db:
                    found_count += 1
                    # 매칭된 행 색칠하기
                    r_item = ref_db[key]
                    for col in range(8, 13):  # 8~12열 (H~L)
                        ws.cell(row=r_item, column=col).fill = light_green_fill

                # 2. 빨간색(SLM참조)인지 확인
                if key in ref_db_color:
                    referenced_count += 1

            # 조건 모드에 따른 결과 판정
            if condition_mode == "OR":
                # 하나라도 존재하면 True
                exist_bool = found_count > 0
                ref_bool = referenced_count > 0
            else:
                # AND (기본값): 모든 페이지가 존재해야 True
                exist_bool = found_count == total_pages
                ref_bool = referenced_count == total_pages

            cell_d.value = "존재" if exist_bool else "미존재"
            cell_e.value = "참조" if ref_bool else "미참조"

        else:
            # 참조 문서나 페이지 정보가 없는 경우
            cell_d.value = "미존재"
            cell_e.value = "미참조"

    # 5. 저장
    try:
        wb.save(output_file)
        print(f"완료! 결과가 '{output_file}'(으)로 저장되었습니다.")
    except Exception as e:
        print(f"저장 중 오류 발생: {e}")


# --- 메인 실행 블록 ---
if __name__ == "__main__":
    # 인자 개수 확인
    # 기본: script.py input output (len=3)
    # 옵션: script.py input output OR (len=4)
    if len(sys.argv) < 3:
        print(
            "사용법: python 스크립트명.py <입력파일경로> <출력파일경로> [조건옵션:AND|OR]"
        )
        print("예시 (기본 AND): python convert.py input.xlsx output.xlsx")
        print("예시 (OR 모드): python convert.py input.xlsx output.xlsx OR")
    else:
        input_path = sys.argv[1]
        output_path = sys.argv[2]

        # 조건 옵션 파싱 (기본값 AND)
        mode_arg = "AND"
        if len(sys.argv) >= 4:
            # 입력받은 값을 대문자로 변환하여 확인
            if sys.argv[3].upper() == "OR":
                mode_arg = "OR"

        process_excel_log(input_path, output_path, mode_arg)
