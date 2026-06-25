#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import sys
import json
import argparse
import datetime
import time
import re
import httpx

# pandas 및 openpyxl 의존성 검사
try:
    import pandas as pd
except ImportError:
    print("🚨 [Error] 'pandas' 라이브러리가 설치되어 있지 않습니다.")
    print("테스트 실행 전에 다음 명령어로 의존성을 설치해 주세요:")
    print("  pip install pandas openpyxl")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("🚨 [Error] 'openpyxl' 라이브러리가 설치되어 있지 않습니다.")
    print("테스트 실행 전에 다음 명령어로 의존성을 설치해 주세요:")
    print("  pip install openpyxl")
    sys.exit(1)

# langchain 의존성 검사
try:
    from langchain_openai import ChatOpenAI
    from langchain_core.messages import HumanMessage
except ImportError:
    print("🚨 [Error] 'langchain-openai' 또는 'langchain-core' 라이브러리가 설치되어 있지 않습니다.")
    print("테스트 실행 전에 다음 명령어로 의존성을 설치해 주세요:")
    print("  pip install langchain-openai langchain-core")
    sys.exit(1)


def parse_goldenset_pages(page_str):
    if not page_str:
        return set()
    pages = set()
    for p in str(page_str).split(','):
        p = p.strip()
        if p.isdigit():
            pages.add(int(p))
    return pages


def extract_graph_was_pages(graph_was_response):
    pages = set()
    if not graph_was_response:
        return pages
    search_result = graph_was_response.get("search_result", [])
    for doc in search_result:
        page_val = doc.get("page")
        if isinstance(page_val, list):
            for p in page_val:
                if isinstance(p, int):
                    pages.add(p)
                elif isinstance(p, str) and p.isdigit():
                    pages.add(int(p))
        elif isinstance(page_val, (int, float)):
            pages.add(int(page_val))
        elif isinstance(page_val, str) and page_val.isdigit():
            pages.add(int(page_val))
    return pages


def extract_answer_pages(final_answer):
    pages = set()
    if not final_answer:
        return pages
    matches = re.findall(r'p\.\s*([0-9\s,]+)', str(final_answer))
    for match in matches:
        for part in match.split(','):
            part = part.strip()
            if part.isdigit():
                pages.add(int(part))
    return pages


def write_summary_sheet(workbook, results):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    if "결과 요약" in workbook.sheetnames:
        del workbook["결과 요약"]
        
    summary_sheet = workbook.create_sheet(title="결과 요약", index=0)
    summary_sheet.views.sheetView[0].showGridLines = True
    
    # 폰트 및 채우기 설정
    title_font = Font(name="맑은 고딕", size=16, bold=True, color="1F497D")
    subtitle_font = Font(name="맑은 고딕", size=10, italic=True, color="595959")
    header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="맑은 고딕", size=11, bold=True)
    normal_font = Font(name="맑은 고딕", size=10)
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    
    # 테두리 설정
    thin_side = Side(style='thin', color='D9D9D9')
    thick_bottom = Side(style='medium', color='1F497D')
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thick_bottom)
    
    # 타이틀
    summary_sheet["A1"] = "📊 Busan AI RAG 골든셋 테스트 결과 요약 (OpenAI)"
    summary_sheet["A1"].font = title_font
    
    summary_sheet["A2"] = f"생성 일시: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    summary_sheet["A2"].font = subtitle_font
    
    # --- 표 1: 정확도 지표 ---
    summary_sheet["A4"] = "평가 구분"
    summary_sheet["B4"] = "대상 건수"
    summary_sheet["C4"] = "성공/매칭 건수"
    summary_sheet["D4"] = "성공률 / 매칭률 (%)"
    
    for col_let in ["A", "B", "C", "D"]:
        cell = summary_sheet[f"{col_let}4"]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border
        
    total_cases = len(results)
    api_success = sum(1 for r in results if r["상태"] == "SUCCESS")
    g_match = sum(1 for r in results if r["Graph-WAS 매칭 여부"] == "Y")
    a_match = sum(1 for r in results if r["최종 답변 매칭 여부"] == "Y")
    
    metrics = [
        ("API 호출 성공률", total_cases, api_success, (api_success / total_cases * 100) if total_cases > 0 else 0),
        ("Graph-WAS 검색 성공률 (쪽번호 일치)", total_cases, g_match, (g_match / total_cases * 100) if total_cases > 0 else 0),
        ("최종 답변 참조 성공률 (쪽번호 일치)", total_cases, a_match, (a_match / total_cases * 100) if total_cases > 0 else 0),
    ]
    
    for idx, (label, tot, succ, rate) in enumerate(metrics, 5):
        summary_sheet[f"A{idx}"] = label
        summary_sheet[f"B{idx}"] = f"{tot}건"
        summary_sheet[f"C{idx}"] = f"{succ}건"
        summary_sheet[f"D{idx}"] = f"{rate:.1f}%"
        
        for col_let in ["A", "B", "C", "D"]:
            cell = summary_sheet[f"{col_let}{idx}"]
            cell.font = bold_font if col_let == "A" else normal_font
            cell.border = cell_border
            if col_let == "A":
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
    # --- 표 2: 소요 시간 지표 ---
    time_start_row = 10
    summary_sheet[f"A{time_start_row}"] = "소요 시간 구분"
    summary_sheet[f"B{time_start_row}"] = "평균 소요 시간 (ms)"
    summary_sheet[f"C{time_start_row}"] = "최대 소요 시간 (ms)"
    summary_sheet[f"D{time_start_row}"] = "최소 소요 시간 (ms)"
    
    for col_let in ["A", "B", "C", "D"]:
        cell = summary_sheet[f"{col_let}{time_start_row}"]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border
        
    def to_float(val):
        try:
            return float(val)
        except (ValueError, TypeError):
            return 0.0
            
    graph_times = [to_float(r.get("Graph-WAS 소요시간(ms)", 0)) for r in results]
    llm_times = [to_float(r.get("LLM 소요시간(ms)", 0)) for r in results]
    total_times = [to_float(r.get("총 소요시간(ms)", 0)) for r in results]
    
    avg_graph = sum(graph_times) / total_cases if total_cases > 0 else 0
    avg_llm = sum(llm_times) / total_cases if total_cases > 0 else 0
    avg_total = sum(total_times) / total_cases if total_cases > 0 else 0
    
    max_graph = max(graph_times) if graph_times else 0
    max_llm = max(llm_times) if llm_times else 0
    max_total = max(total_times) if total_times else 0
    
    min_graph = min(graph_times) if graph_times else 0
    min_llm = min(llm_times) if llm_times else 0
    min_total = min(total_times) if total_times else 0
    
    time_metrics = [
        ("Graph-WAS 검색 소요시간", avg_graph, max_graph, min_graph),
        ("LLM 답변생성 소요시간", avg_llm, max_llm, min_llm),
        ("전체 프로세스 총 소요시간", avg_total, max_total, min_total)
    ]
    
    for idx, (label, avg, mx, mn) in enumerate(time_metrics, time_start_row + 1):
        summary_sheet[f"A{idx}"] = label
        summary_sheet[f"B{idx}"] = f"{avg:.1f} ms"
        summary_sheet[f"C{idx}"] = f"{mx:.1f} ms"
        summary_sheet[f"D{idx}"] = f"{mn:.1f} ms"
        
        for col_let in ["A", "B", "C", "D"]:
            cell = summary_sheet[f"{col_let}{idx}"]
            cell.font = bold_font if col_let == "A" else normal_font
            cell.border = cell_border
            if col_let == "A":
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
    summary_sheet.column_dimensions["A"].width = 38
    summary_sheet.column_dimensions["B"].width = 18
    summary_sheet.column_dimensions["C"].width = 18
    summary_sheet.column_dimensions["D"].width = 24


def parse_arguments():
    parser = argparse.ArgumentParser(description="Busan AI RAG Goldenset CLI Test App (OpenAI-compatible)")
    
    # 기본 경로 설정
    default_goldenset = os.path.join(os.path.dirname(__file__), "goldenset.json")
    default_output_dir = os.path.join(os.path.dirname(__file__), "reports")
    
    parser.add_argument(
        "--goldenset-path",
        type=str,
        default=default_goldenset,
        help=f"골든셋 JSON 파일 경로 (기본값: {default_goldenset})"
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=30,
        help="테스트를 진행할 질문의 최대 개수 (기본값: 30)"
    )
    parser.add_argument(
        "--graph-url",
        type=str,
        default="http://99.1.82.168:38080",
        help="Graph-WAS 서버 주소 (기본값: http://99.1.82.168:38080)"
    )
    parser.add_argument(
        "--api-url",
        type=str,
        default="https://99.1.2.97:10080/v1",
        help="ChatOpenAI 호환 API Base URL (기본값: https://99.1.2.97:10080/v1)"
    )
    parser.add_argument(
        "--model",
        type=str,
        default="HCX-SEED-THINK-14B",
        help="사용할 LLM 모델명 (기본값: HCX-SEED-THINK-14B)"
    )
    parser.add_argument(
        "--api-key",
        type=str,
        default=None,
        help="API Key (기본값: None)"
    )
    parser.add_argument(
        "--max-tokens",
        type=int,
        default=2048,
        help="생성할 최대 토큰 수 (기본값: 2048, 제한하지 않으려면 0 이하로 설정)"
    )
    parser.add_argument(
        "--temperature",
        type=float,
        default=0.0,
        help="생성 온도 (기본값: 0.0. 반복 루프 발생 시 0.2 ~ 0.7 정도로 올려보세요)"
    )
    parser.add_argument(
        "--frequency-penalty",
        type=float,
        default=0.0,
        help="빈도 페널티 (기본값: 0.0. 반복 억제를 위해 0.1 ~ 1.0 사이 값 추천)"
    )
    parser.add_argument(
        "--presence-penalty",
        type=float,
        default=0.0,
        help="존재 페널티 (기본값: 0.0. 새로운 단어/표현 유도를 위해 0.1 ~ 1.0 사이 값 추천)"
    )
    parser.add_argument(
        "--repetition-penalty",
        type=float,
        default=None,
        help="반복 페널티 (기본값: None. vLLM/Ollama 등 지원 백엔드인 경우 1.1 ~ 1.2 등 지정)"
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=1.0,
        help="각 질문 테스트 사이의 대기 시간(초) (기본값: 1.0)"
    )
    parser.add_argument(
        "--output-dir",
        type=str,
        default=default_output_dir,
        help=f"엑셀 리포트가 저장될 디렉토리 (기본값: {default_output_dir})"
    )
    
    return parser.parse_args()


def load_goldenset(file_path: str, limit: int):
    if not os.path.exists(file_path):
        print(f"❌ 골든셋 파일을 찾을 수 없습니다: {file_path}")
        sys.exit(1)
        
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            
        if not isinstance(data, list):
            print("❌ 골든셋 파일의 루트 노드는 JSON 리스트 형태여야 합니다.")
            sys.exit(1)
            
        print(f"ℹ️ 골든셋 파일 로드 완료. 총 {len(data)}개의 데이터 중 상위 {limit}개를 테스트합니다.")
        return data[:limit]
    except Exception as e:
        print(f"❌ 골든셋 파일 로드 중 오류 발생: {e}")
        sys.exit(1)


def main():
    args = parse_arguments()
    
    # 1. 골든셋 로드
    goldenset_data = load_goldenset(args.goldenset_path, args.limit)
    
    # 2. 테스트 결과 리스트 초기화
    results = []
    
    print("\n==================================================")
    print("🚀 부산 AI RAG 골든셋 자동 테스트 시작 (OpenAI-compatible)")
    print(f"🔗 Graph-WAS 주소: {args.graph_url}")
    print(f"🔗 OpenAI API Base 주소: {args.api_url}")
    print(f"🔗 LLM 모델명: {args.model}")
    print(f"⏳ 테스트 딜레이: {args.delay}초")
    print("==================================================\n")
    
    success_count = 0
    fail_count = 0
    start_time_all = time.time()
    
    # 3. 루프 실행
    for idx, item in enumerate(goldenset_data, 1):
        # '정제 질의문'이 없으면 일반 '질의문' 사용
        user_query = item.get("정제 질의문") or item.get("질의문") or ""
        user_query = user_query.strip()
        
        if not user_query:
            print(f"⚠️ [{idx}/{len(goldenset_data)}] 스킵: 유효한 질의문이 없습니다.")
            continue
            
        print(f"📝 [{idx}/{len(goldenset_data)}] 질문: \"{user_query}\"")
        
        # 기본 기록 정보
        graph_was_params = {"query": user_query}
        graph_was_response = {}
        openai_payload = {}
        openai_response = {}
        final_answer = ""
        status = "SUCCESS"
        error_msg = ""
        
        duration_graph_ms = 0.0
        duration_llm_ms = 0.0
        
        try:
            # --- 1단계: Graph-WAS API 호출 ---
            print("  └─ 🔍 [1/2] Graph-WAS 검색 요청 중...")
            t_graph_start = time.time()
            
            with httpx.Client(timeout=30.0) as client:
                graph_res = client.post(
                    f"{args.graph_url}/search",
                    json=graph_was_params
                )
                graph_res.raise_for_status()
                graph_was_response = graph_res.json()
                
            t_graph_end = time.time()
            duration_graph_ms = (t_graph_end - t_graph_start) * 1000.0
            
            prompt = graph_was_response.get("prompt", "")
            
            # --- 2단계: LangChain ChatOpenAI 최종 답변 생성 중 ---
            print("  └─ ⚡ [2/2] LangChain ChatOpenAI 최종 답변 생성 중...")
            
            llm_kwargs = {
                "base_url": args.api_url,
                "model": args.model,
                "temperature": args.temperature,
                "timeout": 90.0,
            }
            if args.api_key:
                llm_kwargs["api_key"] = args.api_key
            if args.max_tokens and args.max_tokens > 0:
                llm_kwargs["max_tokens"] = args.max_tokens
            if args.frequency_penalty != 0.0:
                llm_kwargs["frequency_penalty"] = args.frequency_penalty
            if args.presence_penalty != 0.0:
                llm_kwargs["presence_penalty"] = args.presence_penalty
                
            model_kwargs = {}
            if args.repetition_penalty is not None:
                model_kwargs["repetition_penalty"] = args.repetition_penalty
            if "gpt-oss" in args.model:
                model_kwargs["reasoning_effort"] = "low"
            if model_kwargs:
                llm_kwargs["model_kwargs"] = model_kwargs
                
            t_llm_start = time.time()
            
            llm = ChatOpenAI(**llm_kwargs)
            messages = [HumanMessage(content=prompt if prompt else user_query)]
            llm_response = llm.invoke(messages)
            
            t_llm_end = time.time()
            duration_llm_ms = (t_llm_end - t_llm_start) * 1000.0
            
            final_answer = llm_response.content
            success_count += 1
            print(f"  ✅ 완료 (소요시간: Graph={duration_graph_ms:.1f}ms, LLM={duration_llm_ms:.1f}ms)")
            
            # 리포트 저장을 위한 가짜(Mock) openai_payload, openai_response 구성 (필드 호환성 유지)
            openai_payload = {
                "llm_kwargs": {k: v for k, v in llm_kwargs.items() if k != "api_key"},
                "messages": [m.content for m in messages]
            }
            openai_response = {
                "content": final_answer,
                "response_metadata": getattr(llm_response, "response_metadata", {})
            }
                
        except Exception as e:
            status = "FAIL"
            error_msg = str(e)
            final_answer = f"🚨 통신 오류: {error_msg}"
            fail_count += 1
            print(f"  ❌ 호출 실패: {error_msg}")
            
        # 쪽번호 파싱 및 매칭 판단
        goldenset_page_str = item.get("적용 쪽번호") or ""
        goldenset_pages = parse_goldenset_pages(goldenset_page_str)
        graph_was_pages = extract_graph_was_pages(graph_was_response)
        final_answer_pages = extract_answer_pages(final_answer)

        graph_was_match = "Y" if (goldenset_pages & graph_was_pages) else "N"
        final_answer_match = "Y" if (goldenset_pages & final_answer_pages) else "N"

        # 결과 추가
        results.append({
            "No": idx,
            "사용자 질의문": user_query,
            "Graph-WAS 입력 파라미터": json.dumps(graph_was_params, ensure_ascii=False, indent=2),
            "Graph-WAS 응답값": json.dumps(graph_was_response, ensure_ascii=False, indent=2),
            "LLM 입력 파라미터": json.dumps(openai_payload, ensure_ascii=False, indent=2),
            "LLM 응답값": json.dumps(openai_response, ensure_ascii=False, indent=2),
            "최종 답변": final_answer,
            "상태": status,
            "에러 메시지": error_msg,
            "Graph-WAS 소요시간(ms)": round(duration_graph_ms, 2),
            "LLM 소요시간(ms)": round(duration_llm_ms, 2),
            "총 소요시간(ms)": round(duration_graph_ms + duration_llm_ms, 2),
            "골든셋 정답페이지": goldenset_page_str,
            "Graph-WAS 추출페이지": ", ".join(map(str, sorted(graph_was_pages))),
            "최종 답변 추출페이지": ", ".join(map(str, sorted(final_answer_pages))),
            "Graph-WAS 매칭 여부": graph_was_match,
            "최종 답변 매칭 여부": final_answer_match,
        })
        
        # 딜레이 대기
        if idx < len(goldenset_data) and args.delay > 0:
            time.sleep(args.delay)
            
    # 4. 결과 DataFrame 생성 및 엑셀 저장
    df = pd.DataFrame(results)
    
    # 출력 폴더 생성
    os.makedirs(args.output_dir, exist_ok=True)
    
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"test_report_{timestamp}.xlsx"
    output_path = os.path.join(args.output_dir, file_name)
    
    print("\n💾 엑셀 파일 저장 중...")
    try:
        # Excel 파일 저장 및 스타일링
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="테스트 결과")
            
            # 셀 스타일링을 위한 openpyxl 객체 접근
            workbook = writer.book
            worksheet = writer.sheets["테스트 결과"]
            
            # 결과 요약 시트 작성
            write_summary_sheet(workbook, results)
            
            # 컬럼 너비 자동 조정
            for col in worksheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    # 줄바꿈이 있는 텍스트는 가장 긴 라인을 기준으로 계산
                    val_str = str(cell.value or "")
                    lines = val_str.split("\n")
                    for line in lines:
                        if len(line) > max_len:
                            max_len = len(line)
                # 너무 넓어지지 않도록 최대 너비 제한 (예: 60)
                worksheet.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 60)
                
            # 정렬 및 줄바꿈 허용 (내용이 긴 열은 정렬 방식을 줘서 엑셀에서 보기 편하도록 함)
            from openpyxl.styles import Alignment, PatternFill, Font
            
            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
            
            # 헤더 스타일 적용
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
            # 바디 텍스트 스타일링 (총 17개 컬럼)
            for row in worksheet.iter_rows(min_row=2, max_row=len(results) + 1, min_col=1, max_col=17):
                for cell in row:
                    # No, 상태, 시간, 쪽번호 관련 열은 가운데 정렬
                    col_idx = cell.column
                    if col_idx in [1, 8, 10, 11, 12, 13, 14, 15, 16, 17]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    cell.font = Font(name="맑은 고딕", size=10)
                    
        print(f"🎉 테스트 리포트 저장 완료! -> [보고서 파일]({output_path})")
    except Exception as e:
        print(f"❌ 엑셀 저장 실패: {e}")
        
    # 최종 결과 요약 출력
    duration_all = time.time() - start_time_all
    print("\n==================================================")
    print("📊 테스트 결과 요약")
    print(f"⏱️ 총 소요 시간: {duration_all:.2f}초")
    print(f"✅ 성공: {success_count}건")
    print(f"❌ 실패: {fail_count}건")
    print(f"📈 성공률: {(success_count / len(goldenset_data)) * 100:.1f}%")
    print("==================================================\n")


if __name__ == "__main__":
    main()
