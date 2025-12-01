import sys
import openpyxl
from openpyxl.styles import Font

def normalize_val(val):
    """값을 문자열로 변환하고, 정수형 실수는 정수 문자로 변환 (예: 1.0 -> '1')"""
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()

def parse_pages(page_str):
    """
    페이지 문자열을 파싱하여 개별 페이지 번호의 리스트(문자열)를 반환
    예: "1, 3~5" -> {'1', '3', '4', '5'}
    """
    pages = set()
    s = str(page_str).strip()
    if not s or s.lower() == 'nan' or s == 'none':
        return pages

    # 쉼표로 분리
    parts = s.split(',')
    for part in parts:
        part = part.strip()
        if '~' in part:
            # 범위 처리 (예: 1~3)
            try:
                start_s, end_s = part.split('~')
                start = int(start_s)
                end = int(end_s)
                for i in range(start, end + 1):
                    pages.add(str(i))
            except ValueError:
                pages.add(part)
        else:
            # 단일 페이지
            pages.add(part)
    return pages

def process_excel_log(input_file, output_file):
    print(f"[{input_file}] 파일을 읽는 중...")

    # 1. 엑셀 파일 불러오기
    try:
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active
    except FileNotFoundError:
        print(f"오류: 입력 파일 '{input_file}'을(를) 찾을 수 없습니다.")
        sys.exit(1)
    except Exception as e:
        print(f"오류 발생: {e}")
        sys.exit(1)

    # 2. 기준 데이터 구축 (I열: file, J열: page_num)
    #    열 인덱스(0-based): I=8, J=9
    ref_db = set()
    
    # 데이터 읽기 최적화를 위해 values_only 사용
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= 10:
            file_val = normalize_val(row[8])
            page_val = normalize_val(row[9])
            
            if file_val and page_val:
                ref_db.add((file_val, page_val))

    print("데이터 분석 및 처리 중...")

    # 3. D열, E열 신규 삽입 (4번째 위치에 2개 열)
    ws.insert_cols(4, 2)
    
    # 헤더 설정
    ws['D1'] = "참조문서검색"
    ws['E1'] = "SLM참조여부"

    # 4. 데이터 처리 (행별 반복)
    #    삽입 후 열 인덱스: B=1, C=2, D=3, E=4
    for row in ws.iter_rows(min_row=2):
        cell_b = row[1]  # 답변참조문서
        cell_c = row[2]  # 참조페이지
        cell_d = row[3]  # 참조문서검색 (Target)
        cell_e = row[4]  # SLM참조여부 (Target)
        
        # --- [Logic 1] 참조문서검색 ---
        b_val = normalize_val(cell_b.value)
        c_pages = parse_pages(cell_c.value)
        
        # 조건: 파일명이 있고, 페이지가 하나 이상 있으며, 모든 페이지가 DB에 존재해야 함
        if b_val and c_pages:
            all_exist = True
            for page in c_pages:
                if (b_val, page) not in ref_db:
                    all_exist = False
                    break
            cell_d.value = "존재" if all_exist else "미존재"
        else:
            cell_d.value = "미존재"

        # --- [Logic 2] SLM참조여부 (빨간색 확인) ---
        is_referenced = False
        
        def check_red(cell):
            # 글자색이 있고 RGB값이 빨간색(FFFF0000 등)인지 확인
            if cell.font and cell.font.color and hasattr(cell.font.color, 'rgb'):
                color = str(cell.font.color.rgb)
                # Alpha channel 포함(FFFF0000) 또는 미포함(FF0000/00FF0000) 케이스 대응
                if 'FF0000' in color and color != '00000000': 
                    # 더 정확하게는 보통 'FFFF0000' 입니다.
                    if color == 'FFFF0000' or color == '00FF0000':
                        return True
            return False

        if check_red(cell_b) or check_red(cell_c):
            is_referenced = True
            
        cell_e.value = "참조" if is_referenced else "미참조"

    # 5. 저장
    try:
        wb.save(output_file)
        print(f"완료! 결과가 '{output_file}'(으)로 저장되었습니다.")
    except Exception as e:
        print(f"저장 중 오류 발생: {e}")

# --- 메인 실행 블록 ---
if __name__ == "__main__":
    # 인자 개수 확인 (스크립트명 + 입력파일 + 출력파일 = 3개)
    if len(sys.argv) != 3:
        print("사용법: python 스크립트명.py <입력파일경로> <출력파일경로>")
        print("예시: python convert.py input.xlsx output.xlsx")
    else:
        input_path = sys.argv[1]
        output_path = sys.argv[2]
        process_excel_log(input_path, output_path)